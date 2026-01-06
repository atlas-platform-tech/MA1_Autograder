# writers/build_instructor_master_workbook.py

import os
from typing import List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utilities.paths import ws_path  # ✅ workspace-aware template path


SUMMARY_SHEET_NAME = "Summary_Template"

SUMMARY_HEADER_ROW = 4
SUMMARY_START_ROW = 5

# Summary columns
COL_NAME = "A"
COL_AUTO_TOTAL = "B"
COL_MANUAL_ADJ = "C"
COL_FINAL_TOTAL = "D"
COL_INCOME_TOTAL = "E"
COL_UNIT_TOTAL = "F"
COL_CURRENCY_TOTAL = "G"

# Grading sheet anchors (in each student's grading sheet)
ANCHOR_INCOME = "$F$8"
ANCHOR_UNIT = "$F$14"
ANCHOR_CURRENCY = "$F$23"
ANCHOR_OVERALL = "$F$24"


def _grade_files_in_folder(graded_path: str) -> List[str]:
    graded_path = os.path.abspath(graded_path)
    files = []
    for fn in os.listdir(graded_path):
        if fn.lower().endswith("_ma1_grade.xlsx"):
            files.append(os.path.join(graded_path, fn))
    files.sort()
    return files


def _xl_escape_path(p: str) -> str:
    return os.path.abspath(p).replace("/", "\\")


def _external_link_formula_local(grade_filename: str, cell_ref: str) -> str:
    """
    ✅ Relative link (BEST) because INSTRUCTOR_MASTER.xlsx is saved in the SAME folder
    as all *_MA1_Grade.xlsx files.

    Example:
      ='[Adrian_Torres_MA1_Grade.xlsx]Grading Sheet'!$F$24

    This avoids absolute-path link errors and "Manage Workbook Links" headaches.
    """
    return f"='[{grade_filename}]Grading Sheet'!{cell_ref}"


def _ensure_summary_headers(ws: Worksheet) -> None:
    """
    Ensure the section columns have labels (in case template is blank there).
    """
    headers = {
        COL_INCOME_TOTAL: "Income Total",
        COL_UNIT_TOTAL: "Unit Total",
        COL_CURRENCY_TOTAL: "Currency Total",
    }
    for col, label in headers.items():
        cell = f"{col}{SUMMARY_HEADER_ROW}"
        if ws[cell].value in (None, ""):
            ws[cell].value = label


def build_instructor_master_workbook(
    graded_path: str,
    template_path: str | None = None,
    output_filename: str = "INSTRUCTOR_MASTER.xlsx",
) -> str:
    """
    Builds a summary-only instructor master workbook.

    - Uses Template_Master.xlsx -> Summary_Template sheet only
    - Writes one row per student grade file
    - Links to each student's grading sheet totals & section totals

    Saves into:
      graded_output/<course>/INSTRUCTOR_MASTER.xlsx

    Returns: absolute path to created workbook
    """
    graded_path = os.path.abspath(graded_path)

    if not os.path.isdir(graded_path):
        raise FileNotFoundError(f"graded_path not found: {graded_path}")

    # ✅ Default template location is workspace templates folder
    if template_path is None:
        template_path = ws_path("templates", "Template_Master.xlsx")
    else:
        template_path = os.path.abspath(template_path)

    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"❌ Template_Master.xlsx not found at:\n{template_path}\n\n"
            f"Place it here:\nDocuments/MA1_Autograder/templates/Template_Master.xlsx"
        )

    grade_files = _grade_files_in_folder(graded_path)
    if not grade_files:
        raise FileNotFoundError(f"No *_MA1_Grade.xlsx files found in: {graded_path}")

    wb = load_workbook(template_path)

    if SUMMARY_SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Template missing sheet: {SUMMARY_SHEET_NAME}")

    ws_summary = wb[SUMMARY_SHEET_NAME]
    _ensure_summary_headers(ws_summary)

    # Help Excel recalc links on open
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    row = SUMMARY_START_ROW

    for full_grade_path in grade_files:
        grade_filename = os.path.basename(full_grade_path)
        student_base = grade_filename.replace("_MA1_Grade.xlsx", "")

        # Name + hyperlink to the grade file
        name_cell = f"{COL_NAME}{row}"
        ws_summary[name_cell].value = student_base
        ws_summary[name_cell].hyperlink = _xl_escape_path(full_grade_path)

        # Auto Total (F24)
        ws_summary[f"{COL_AUTO_TOTAL}{row}"].value = _external_link_formula_local(
            grade_filename, ANCHOR_OVERALL
        )

        # Manual Adj default 0
        manual_cell = f"{COL_MANUAL_ADJ}{row}"
        if ws_summary[manual_cell].value in (None, ""):
            ws_summary[manual_cell].value = 0

        # Final Total = Auto + Manual
        ws_summary[f"{COL_FINAL_TOTAL}{row}"].value = f"={COL_AUTO_TOTAL}{row}+{COL_MANUAL_ADJ}{row}"

        # Section totals
        ws_summary[f"{COL_INCOME_TOTAL}{row}"].value = _external_link_formula_local(
            grade_filename, ANCHOR_INCOME
        )
        ws_summary[f"{COL_UNIT_TOTAL}{row}"].value = _external_link_formula_local(
            grade_filename, ANCHOR_UNIT
        )
        ws_summary[f"{COL_CURRENCY_TOTAL}{row}"].value = _external_link_formula_local(
            grade_filename, ANCHOR_CURRENCY
        )

        row += 1

    out_path = os.path.join(graded_path, output_filename)
    wb.save(out_path)
    return out_path

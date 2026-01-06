# orchestrator/phase1_grade_all.py

import os
from openpyxl import load_workbook

from graders.income_analysis.grade_income_analysis import grade_income_analysis
from writers.write_income_analysis_scores import write_income_analysis_scores

# ---- Unit Conversions V2 imports ----
from graders.unit_conversions.unit_conversions_checker_v2 import grade_unit_conversions_tab_v2
from writers.unit_conversions_writer_v2 import write_unit_conversions_scores_v2
# -----------------------------------

# ---- Currency Conversion V2 imports ----
from graders.currency_conversion.grade_currency_conversion_tab_v2 import grade_currency_conversion_tab_v2
from writers.write_currency_conversion_results_v2 import write_currency_conversion_results_v2
# ---------------------------------------


def phase1_grade_all_students(submissions_path, graded_output_path):
    """
    Grades the formula-based parts of every student's MA1 workbook.
    (Chart export and insertion happen in later phases.)
    """

    print("\n📘 PHASE 1 — Grading all students...\n")

    for filename in os.listdir(submissions_path):
        if not filename.endswith(".xlsx"):
            continue

        student_name = filename.replace("_MA1.xlsx", "")
        submission_file = os.path.join(submissions_path, filename)
        grading_file = os.path.join(graded_output_path, f"{student_name}_MA1_Grade.xlsx")

        try:
            student_wb = load_workbook(submission_file, data_only=False)
            grading_wb = load_workbook(grading_file)

            ws_income = student_wb["Income Analysis"]
            ws_grading = grading_wb["Grading Sheet"]

            # -----------------------------
            # INCOME ANALYSIS
            # -----------------------------
            ia_results = grade_income_analysis(ws_income)
            write_income_analysis_scores(ws_grading, ia_results)

            # -----------------------------
            # UNIT CONVERSIONS — V2 ONLY
            # -----------------------------
            try:
                ws_unit = student_wb["Unit Conversions"]
                uc_results = grade_unit_conversions_tab_v2(ws_unit)
                write_unit_conversions_scores_v2(ws_grading, uc_results)
            except Exception as e:
                print(f"⚠️ Unit Conversions error for {student_name}: {e}")

            # -----------------------------
            # CURRENCY CONVERSION — V2 ONLY
            # -----------------------------
            try:
                ws_currency = student_wb["Currency Conversion"]
                cc_results = grade_currency_conversion_tab_v2(ws_currency, student_name)
                write_currency_conversion_results_v2(ws_grading, cc_results)
            except Exception as e:
                print(f"⚠️ Currency Conversion error for {student_name}: {e}")

            grading_wb.save(grading_file)
            print(f"✅ Graded: {student_name}")

        except Exception as e:
            print(f"❌ Error grading {student_name}: {e}")

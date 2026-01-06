# orchestrator/grade_single.py

import os
from openpyxl import load_workbook

from graders.income_analysis.grade_income_analysis import grade_income_analysis
from writers.write_income_analysis_scores import write_income_analysis_scores

# ---- Unit Conversions V2 ----
from graders.unit_conversions.unit_conversions_checker_v2 import grade_unit_conversions_tab_v2
from writers.unit_conversions_writer_v2 import write_unit_conversions_scores_v2
# -----------------------------

# ---- Currency Conversion V2 ----
from graders.currency_conversion.grade_currency_conversion_tab_v2 import grade_currency_conversion_tab_v2
from writers.write_currency_conversion_results_v2 import write_currency_conversion_results_v2
# --------------------------------


def grade_single_file(submission_path: str, graded_output_folder: str) -> dict:
    """
    Grade a single student's MA1 workbook.

    Assumes:
      - submission_path points to:
            student_submissions/<course_label>/First_Last_MA1.xlsx
      - A matching grading sheet already exists:
            graded_output/<course_label>/First_Last_MA1_Grade.xlsx
    """

    if not os.path.exists(submission_path):
        raise FileNotFoundError(f"Submission file not found: {submission_path}")

    if not os.path.isdir(graded_output_folder):
        raise FileNotFoundError(f"Graded output folder not found: {graded_output_folder}")

    # --- Infer grading sheet filename ---
    submission_name = os.path.basename(submission_path)
    base_name, ext = os.path.splitext(submission_name)
    grading_filename = f"{base_name}_Grade{ext}"
    grading_path = os.path.join(graded_output_folder, grading_filename)

    if not os.path.exists(grading_path):
        raise FileNotFoundError(f"Grading sheet not found: {grading_path}")

    # --- Load workbooks ---
    student_wb = load_workbook(submission_path, data_only=False)
    grading_wb = load_workbook(grading_path)

    # --- Grading sheet ---
    try:
        grading_ws = grading_wb["Grading Sheet"]
    except KeyError:
        raise KeyError("The 'Grading Sheet' tab was not found in the grading template workbook.")

    results_out = {}

    # ------------------------------
    # INCOME ANALYSIS
    # ------------------------------
    try:
        ws_income = student_wb["Income Analysis"]
        ia_results = grade_income_analysis(ws_income)
        write_income_analysis_scores(grading_ws, ia_results)
        results_out["income_analysis"] = ia_results
    except Exception as e:
        print(f"⚠️ Income Analysis error: {e}")

    # ------------------------------
    # UNIT CONVERSIONS — V2
    # ------------------------------
    try:
        ws_unit = student_wb["Unit Conversions"]
        uc_results = grade_unit_conversions_tab_v2(ws_unit)
        write_unit_conversions_scores_v2(grading_ws, uc_results)
        results_out["unit_conversions_v2"] = uc_results
    except Exception as e:
        print(f"⚠️ Unit Conversions error: {e}")

    # ------------------------------
    # CURRENCY CONVERSION — V2
    # Name source: filename "First_Last_MA1.xlsx" -> "First_Last"
    # ------------------------------
    try:
        student_name = base_name.replace("_MA1", "")  # "First_Last"
        ws_currency = student_wb["Currency Conversion"]
        cc_results = grade_currency_conversion_tab_v2(ws_currency, student_name)
        write_currency_conversion_results_v2(grading_ws, cc_results)
        results_out["currency_conversion_v2"] = cc_results
    except Exception as e:
        print(f"⚠️ Currency Conversion error: {e}")

    # Save the updated grading workbook
    grading_wb.save(grading_path)

    return results_out
